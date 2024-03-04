import openpyxl
import os
import numpy as np
from itertools import chain
from openpyxl import Workbook



# fill in some columns that are unique for blocks in one attempt
def fill_in(ws1, ws2, start_r, end_r, ref_r, attempt,KCmissing):
    #copy sid(2)>student_id(1), timestamp(3) > 2, div_id(6) > 8
    for r in range(start_r, end_r+len(KCmissing)):
        sid = ws1.cell(row = ref_r, column = 2)
        ws2.cell(row = r, column = 1).value = sid.value
        time = ws1.cell(row = ref_r, column = 3)
        ws2.cell(row = r, column = 2).value = time.value
        probname = ws1.cell(row = ref_r, column = 6)
        ws2.cell(row = r, column = 9).value = probname.value
        ws2.cell(row = r, column = 3).value ="ATTEMPT"
        ws2.cell(row = r, column = 4).value ="Parsons problems"
        ws2.cell(row = r, column = 5).value = "programming"
        ws2.cell(row = r, column = 7).value = attempt
        ws2.cell(row = r, column = 13).value = 1
        # check if it is last attempt by checking problem name with next same users' action
        ws2.cell(row = r, column = 8).value = 0
        nxt = 1
        while ws1.cell(row = ref_r+nxt, column = 2).value != ws1.cell(row = ref_r, column = 2).value:
            nxt = nxt + 1
        if(ws1.cell(row = ref_r+nxt, column = 6).value != ws1.cell(row = ref_r, column = 6).value):
            ws2.cell(row = r, column = 8).value = 1

        # if this is the last attempt, mark all missing KCs as incorrect
        if r in range(end_r, end_r + len(KCmissing)):
            ws2.cell(row = r, column = 10).value = "INCORRECT"
            KC, block = (KCmissing[r-end_r]).split("|")
            ws2.cell(row = r, column = 11).value = "missing"+block
            ws2.cell(row = r, column = 6).value = "missing "+block +" in last attempt"
            ws2.cell(row = r, column = 12).value = KC
            
            

    
# define the dictionary that stores the answer, distractors, and KC info
def collect_ans(ws1,AnsDict):

    #answer
    (AnsDict['exp1_q5_pp'])[0] = ['0_0','1_1','2_1','3_4_2','7_2','9_2','10_1']
    (AnsDict['exp1_pp1a'])[0] = ['0_0','1_1','3_1','4_2','6_3','8_2','9_1']
    (AnsDict['Count_Target_In_Range_Order'])[0] = ['0_0','1_1','3_1','5_2','7_2','9_3','11_1']
    (AnsDict['Total_Dict_Values_PP'])[0] = ['0_0','2_1','3_1','5_2','7_1']
    (AnsDict['exp1_pp3'])[0] = ['0_0','2_1','3_1','4_1']
    
    # distractors
    (AnsDict['exp1_q5_pp'])[1] = ['5','6','8']
    (AnsDict['exp1_pp1a'])[1] = ['2','5','7']
    (AnsDict['Count_Target_In_Range_Order'])[1] = ['2', '4', '6', '8','10']
    (AnsDict['Total_Dict_Values_PP'])[1] = ['1','4', '6']
    (AnsDict['exp1_pp3'])[1] = ['1','5']

    #KC 
    #Attribute: 
    (AnsDict['exp1_q5_pp'])[2] = ['FunctionDef,args','Assign','For','get() function,Attribute,args','Assign,BinOp,Add','append() function,Attribute,args','Return']
    (AnsDict['exp1_pp1a'])[2] = ['FunctionDef,args','Assign','While,Comparison Operator < Less than,len() function,args','If Statement,BoolOp,And,Comparison Operator == Equal,Subscript','Return','Assignment Operator +=','Return']
    (AnsDict['Count_Target_In_Range_Order'])[2] = ['FunctionDef,args','Assign','For,range() function,args,BinOp,Add','Assign,Subscript','Comparison Operator == Equal','Assign,BinOp,Add','Return']
    (AnsDict['Total_Dict_Values_PP'])[2] = ['FunctionDef,args','Assign','For','Assignment Operator +=,Subscript','Return']
    (AnsDict['exp1_pp3'])[2] = ['FunctionDef,args','Assign,max() function,args','Assign,max() function,args','Return,BinOp,Sub']

    # distractor KC
    (AnsDict['exp1_q5_pp'])[3] = ['get() function,Attribute,args','get() function,Attribute,args','Assign,BinOp,Add']
    (AnsDict['exp1_pp1a'])[3] = ['Assign','If Statement,BoolOp,And,Comparison Operator == Equal,Subscript','Return']
    (AnsDict['Count_Target_In_Range_Order'])[3] = ['Assign', 'For,range() function,args,BinOp,Add', 'Assign,Subscript', 'Comparison Operator == Equal','Assign,BinOp,Add']
    (AnsDict['Total_Dict_Values_PP'])[3] = ['FunctionDef,args','For', 'Assignment Operator +=, Subscript']
    (AnsDict['exp1_pp3'])[3] = ['FunctionDef,args','Return,BinOp,Sub']
    
    return AnsDict

    
    
# compare the current and last attempt to find the blocks that are changed or added
def find_changes(last, curr, probname, line):
    changes = []
    deleted = []
    res = []
    if curr == []:
        return []
    if last[0] == '':
        return curr
    curr_tmp = []
    curr_blocks = []
    # 1. find new added elements
    for item in curr:
        if item not in last:
            changes.append(item)
        curr_tmp.append(item)
        curr_blocks.append((item.split('_'))[0])
        

    # fill in places for removed blocks so to compare change in order for existing blocks
    while len(curr_tmp) < len(last):
        for i in range(0, len(last)):
            if ((last[i]).split('_'))[0] not in curr_blocks:
                curr_tmp.insert(i, 'block removed')
                
    #  find elements that changed order
    for block in curr_tmp:
        if block != 'block removed' and block in last and (curr_tmp.index(block) != last.index(block)):
            changes.append(block)
    # iterate through curr, if the item is in the change list, put them in return change in order
    for item in curr:
        if item in changes:
            res.append(item)
    return res
    


#find attempt num
def get_attempt(c,line,ws1,r, ws2):
    res = c.split('-')
    apt = ''
    attempt = 0
    #fetch attempt num from the raw data
    for letter in res:
        if letter[0] == 'c':
            apt = letter[1:len(letter)]
    attempt = int(apt)

    
    if attempt == 1:
        lx = line - 1
        #check if there's other attempt (same student same problem) before
        while (str(ws1.cell(row = lx, column = 2).value) != str(ws1.cell(row = line, column = 2).value)) or (str(ws1.cell(row = lx, column = 6).value) != str(ws1.cell(row = line, column = 6).value)) or (str(ws1.cell(row = lx, column = 4).value) != 'parsonsMove' and str(ws1.cell(row = lx, column = 4).value) != 'parsons'):
            lx = lx - 1
            if lx <= 1:
                return attempt
        content_prev = ws1.cell(row = lx, column = 5).value
        typ2,le2,ri2,c2 = content_prev.split('|')
        res2 = c2.split('-')
        if 's' not in res2:
            return attempt

    #if not 1, find the previous attempt and accumulate from previous filled attempt number
    l = r - 1
    while (str(ws2.cell(row = l, column = 1).value) != str(ws1.cell(row = line, column = 2).value)) or (str(ws2.cell(row = l, column = 9).value) != str(ws1.cell(row = line, column = 6).value)):
        l = l-1
        if l <= 1:
            attempt = 1;
            return attempt
    attempt_prev = ws2.cell(row = l, column = 7).value
    attempt = int(attempt_prev) + 1
    return attempt
        
    

# if the input contains distractors which have nothing to correspond to in answer, add fill-in strings in answer list to avoid seg fault
def add_distractor(ans, curr, distractor):
    res = []
    for item in ans:
        res.append(item)
    for block in curr:
        bkstart = (block.split('_'))[0]
        if bkstart in distractor:
            # insert at that index of curr_ans_tmp
            if distractor_mode == True:
                addinto = "distractor " + str(bkstart)
                res.insert(curr.index(block), addinto)
            else:
                addinto = "extra " + str(bkstart)
                res.append(addinto)
            
    return res


# find the corresponding KC of a block (without indentation) in a certain problem
def find_corres_KCs(blockNum,probname,AnsDict):
    answer = AnsDict[probname][0]
    distractors = AnsDict[probname][1]
    KClist = AnsDict[probname][2]
    KC_disList = AnsDict[probname][3]
    if len(blockNum) == 1:
        for item in answer:
            b = (item.split("_"))[0]
            if b == blockNum[0]:
                i = answer.index(item)
                KC = KClist[i]
                return KC
    else:
        res = ""
        for block in blockNum:
            for item in answer:
                b = (item.split("_"))[0]
                if b == block:
                    i = answer.index(item)
                    KC = KClist[i]
                    if res == "":
                        res = KC
                    else:
                        res = res + "," + KC
        return res



# modify the KC list to let it fit the input
def check_correct_KC(probname, inp, AnsDict):
    res = []
    answer = AnsDict[probname][0]
    KC = AnsDict[probname][2]
    #if input is shorter, there would be something like "6_8_0" which is correct, but the answer recorded may be "6_0""8_0"
    #so need to concatenate the KC of both blocks to the one block 6_8_0 for this attempt
    if len(inp) < len(answer):
        for b in inp:
            blocks = (b.split("_"))[:-1]
            if probname == "exp1_q5_pp" and blocks[0] == "3" and blocks[1] == "4":
                res.append('get() function,Attribute,args')
            else:
                KClist = find_corres_KCs(blocks, probname, AnsDict)
                res.append(KClist)
    else:
        for item in KC:
            res.append(item)
    return res
                
                
                
        
# if the input contain distractor, add KC for the distractors
def add_KC(initial, ans_tmp, KC_distractor, distractorlist):
    res = []
    check_use = []
    for item in initial:
        res.append(item)           
    for i in range(0, len(ans_tmp)):
        string = ans_tmp[i]
        if string[0] == 'd':
            distractor = (string.split())[1]
            inx = distractorlist.index(distractor)
            res.insert(ans_tmp.index(string), KC_distractor[inx])
        if string[0] == 'e':
            distractor = (string.split())[1]
            inx = distractorlist.index(distractor)
            res.append(KC_distractor[inx])
                        
    return res
            

# split the KC string to multiple rows so each row only contain one KC
def split_KCs(start_r, end_r, ws2, KClist,dist):
    for i in range(start_r, end_r):
        ws2.cell(row = i, column = 10).value = ws2.cell(row = start_r, column = 10).value
        ws2.cell(row = i, column = 11).value = ws2.cell(row = start_r, column = 11).value
        if dist:
            ws2.cell(row = i, column = 6).value = "distractor " + str(ws2.cell(row = start_r, column = 11).value)
        else:
            ws2.cell(row = i, column = 6).value = KClist[i-start_r]
        ws2.cell(row = i, column = 12).value = KClist[i-start_r]



    
def data(filename, output, AnsDict, ws1, ws2, out, filt, dismode=True, distVerb=False):
    global distractor_mode
    distractor_mode = dismode
    global last_attempt;
    distV = distVerb


    # annotate the title of each column
    mr = ws1.max_row
    mc = ws1.max_column
    ws2["A1"] = "Anon Student Id" 
    ws2["B1"] = "Time"
    ws2["C1"] = "Student Response Type" 
    ws2["D1"] = "Student Response Subtype"
    ws2["E1"] = "Level(default)"
    ws2["F1"] = "Step Name"
    ws2["G1"] = "Attempt At Step"
    ws2["H1"] = "Is Last Attempt"   
    ws2["I1"] = "Problem Name"
    ws2["J1"] = "Outcome"  #correct / incorrect
    ws2["K1"] = "Input"  #?the block dragged in right
    ws2["L1"] = "KC(Tokens)" #make a list to correspond for each step?
    ws2["M1"] = "Session Id"
    ws2["N1"] = "reference line"
    ws2["O1"] = "should be"

    # initiate for the loop            
    r = 2;
    last_attempt = []
    correct = False;

        

    
    for i in range(2, mr+1):
        # select only cells that are submission attempts and problem names are in the range entered
         if(str(ws1.cell(row = i, column = 4).value) == 'parsons') and ((str(ws1.cell(row = i, column = 6).value) in filt)):

             
            probname = ws1.cell(row = i, column = 6).value
            content = ws1.cell(row = i, column = 5).value
            typ,le,ri,c = content.split('|')

            #get attempt number
            attempt = 0;
            attempt = get_attempt(c,i,ws1,r, ws2)
            ws2.cell(row = r, column = 7).value = attempt



            nxt = 1
            check_last = False
            while ws1.cell(row = i+nxt, column = 2).value != ws1.cell(row = i, column = 2).value:
                nxt = nxt + 1
            if(ws1.cell(row = i+nxt, column = 6).value != ws1.cell(row = i, column = 6).value):
                check_last = True
            
            #check correctness
            if typ == 'incorrect':
                ws2.cell(row = r, column = 10).value = 'INCORRECT'
                ws2.cell(row = r, column = 14).value = i
                correct = False
            
            if typ == 'correct':
                ws2.cell(row = r, column = 10).value = 'CORRECT'
                correct = True
                ws2.cell(row = r, column = 14).value = i
                


            #look into the input result
            inp = ri.split('-')
            if ri == "-":
                inp = []
            correct_ans = AnsDict[probname][0]
            correct_ans_tmp = []
            distractors = AnsDict[probname][1]
            KC = AnsDict[probname][2]
            KC_distractor = AnsDict[probname][3]
            KC_tmp = []

            # modify the reference solution list and corresponding KC list
            correct_ans_tmp = add_distractor(correct_ans, inp, (AnsDict[probname])[1])
            if correct:
                KC_tmp = check_correct_KC(probname, inp, AnsDict)
            else:
                KC_tmp = add_KC(KC,correct_ans_tmp,KC_distractor, distractors)
            KC_missing = []
            check_use = []
            check_ref = []

            # if it's last attempt, check if there's missing KCs
            if check_last:
                for item in inp:
                    Ax = item.split("_")
                    check_use = check_use + Ax[:-1]
                for A in correct_ans:
                    Ab = A.split("_")
                    check_ref = Ab[:-1]
                    for it in check_ref:
                        if it not in check_use:
                            idx = correct_ans.index(A)
                            addList = ((AnsDict[probname][2])[idx]).split(",")
                            for items in addList:
                                items = items+"|"+str(A)
                                KC_missing.append(items)
            if correct:
                KC_missing = []

    



            count_change = 0
            #when it's first submission
            if(last_attempt == []):
                if inp == []:
                    count_change = 1
                    ws2.cell(row = r, column = 6).value = probname + " empty/unchanged in attempt " + str(attempt)
                    r = r + 1
                
                for x in range(0, len(inp)):
                    #if the raw data is cited as correct, then mark all input blocks as correct
                    if(correct):
                        ws2.cell(row = r, column = 10).value = 'CORRECT'
                        ws2.cell(row = r, column = 11).value = inp[x] 
                        KClist = ((KC_tmp[x]).split(","))
                        count_change = count_change + len(KClist)
                        split_KCs(r, r + len(KClist), ws2, KClist, distV)
                        r = r + len(KClist)
                    else: #if the raw data marked incorrect, check each code blocks and find those incorrect 
                        #if the block is corresponding to d-distractor or e-extra in the temporary answer, mark it incorrect
                        if (inp[x] != correct_ans_tmp[x]) and (((correct_ans_tmp[x])[0] == 'd') or ((correct_ans_tmp[x])[0] == 'e')):
                            ws2.cell(row = r, column = 10).value = 'INCORRECT'
                            ws2.cell(row = r, column = 11).value = inp[x]
                            ws2.cell(row = r, column = 15).value = correct_ans_tmp[x]
                            KClist = ((KC_tmp[x]).split(","))
                            count_change = count_change + len(KClist)
                            split_KCs(r, r + len(KClist), ws2, KClist,distV)
                            r = r + len(KClist)
                        if (inp[x] != correct_ans_tmp[x]) and ((correct_ans_tmp[x])[0] != 'd') and ((correct_ans_tmp[x])[0] != 'e'):
                            # if not a distractor, check each block individually
                            if (inp[x] in correct_ans) and (correct_ans.index(inp[x]) == x):
                                ws2.cell(row = r, column = 10).value = 'CORRECT'
                                ws2.cell(row = r, column = 11).value = inp[x]
                                KClist = ((KC[x]).split(","))
                                count_change = count_change + len(KClist)
                                split_KCs(r, r + len(KClist), ws2, KClist, distV)
                                r = r + len(KClist)
                            else:
                                ws2.cell(row = r, column = 10).value = 'INCORRECT'
                                ws2.cell(row = r, column = 11).value = inp[x]
                                ws2.cell(row = r, column = 15).value = correct_ans_tmp[x]
                                KClist = ((KC_tmp[x]).split(","))
                                count_change = count_change + len(KClist)
                                split_KCs(r, r + len(KClist), ws2, KClist,distV)
                                r = r + len(KClist)
        
                        #if the distractor is an added extra block to the answer(relative order has higher priority)
                        if (inp[x] == correct_ans_tmp[x]):
                            ws2.cell(row = r, column = 10).value = 'CORRECT'
                            ws2.cell(row = r, column = 11).value = inp[x]
                            ws2.cell(row = r, column = 6).value = KC_tmp[x]
                            ws2.cell(row = r, column = 12).value = KC_tmp[x]
                            KClist = ((KC_tmp[x]).split(","))
                            count_change = count_change + len(KClist)
                            split_KCs(r, r + len(KClist), ws2, KClist, distV)
                            r = r + len(KClist)


            #if not first attempt, compare with the input from last attempt to find the different ones, check their correctness
            #then act as the changed blocks are inputs, other standards are same as testing the first attempt above
            if last_attempt != []:
                changed = find_changes(last_attempt, inp, probname, i)
                count_change = 0
                if changed == []:
                    count_change = 1
                    ws2.cell(row = r, column = 6).value = probname + " empty/unchanged in attempt " + str(attempt)
                    r = r + 1

                for y in range(0, len(changed)):
                    idx = inp.index(changed[y])
                    if(correct):
                        ws2.cell(row = r, column = 10).value = 'CORRECT'
                        ws2.cell(row = r, column = 11).value = inp[idx]
                        KClist = ((KC_tmp[idx]).split(","))
                        blocks = (inp[idx].split("_"))[:-1]
                        count_change = count_change + len(KClist)
                        split_KCs(r, r + len(KClist), ws2, KClist, distV)
                        r = r + len(KClist)
                    else:
                        if (inp[idx] != correct_ans_tmp[idx]) and (((correct_ans_tmp[idx])[0] == 'd')or ((correct_ans_tmp[idx])[0] == 'e')):
                            ws2.cell(row = r, column = 10).value = 'INCORRECT'
                            ws2.cell(row = r, column = 11).value = inp[idx]
                            ws2.cell(row = r, column = 15).value = correct_ans_tmp[idx]
                            KClist = ((KC_tmp[idx]).split(","))
                            count_change = count_change + len(KClist)
                            split_KCs(r, r + len(KClist), ws2, KClist,distV)
                            r = r + len(KClist)
                        
                        if (inp[idx] != correct_ans_tmp[idx]) and ((correct_ans_tmp[idx])[0] != 'd') and ((correct_ans_tmp[idx])[0] != 'e'):
                            if (inp[idx] in correct_ans) and (correct_ans.index(inp[idx]) == idx):
                                ws2.cell(row = r, column = 10).value = 'CORRECT'
                                ws2.cell(row = r, column = 11).value = inp[idx]
                                KClist = ((KC[idx]).split(","))
                                count_change = count_change + len(KClist)
                                split_KCs(r, r + len(KClist), ws2, KClist,distV)
                                r = r + len(KClist)
                            

                            else:
                                ws2.cell(row = r, column = 10).value = 'INCORRECT'
                                ws2.cell(row = r, column = 11).value = inp[idx]
                                ws2.cell(row = r, column = 15).value = correct_ans_tmp[idx]
                                KClist = ((KC_tmp[idx]).split(","))
                                count_change = count_change + len(KClist)
                                split_KCs(r, r + len(KClist), ws2, KClist,distV)
                                r = r + len(KClist)

                        if (inp[idx] == correct_ans_tmp[idx]):
                            ws2.cell(row = r, column = 10).value = 'CORRECT'
                            ws2.cell(row = r, column = 11).value = inp[idx]
                            KClist = ((KC_tmp[idx]).split(","))
                            count_change = count_change + len(KClist)
                            split_KCs(r, r + len(KClist), ws2, KClist,distV)
                            r = r + len(KClist)
            #end analyzing an attempt. increase the current line by the number of lines filled with content
            #and fill in the other columns needed
            last_attempt = inp;
            start_r = r - count_change
            curr_r = r
            fill_in(ws1, ws2, start_r, r, i, attempt, KC_missing)
            r = curr_r + len(KC_missing)
            #if this is the last attempt of the student in this problem, reset the last_attempt to empty
            nxt = 1
            while ws1.cell(row = i+nxt, column = 2).value != ws1.cell(row = i, column = 2).value:
                nxt = nxt + 1
            if(ws1.cell(row = i+nxt, column = 6).value != ws1.cell(row = i, column = 6).value):
                last_attempt = []


            

    
    out.save(str(output))



# initiate with the parameters entered. then call the data function
def init(filename, output, filt=[], dismode=True, distV=False):
    source = openpyxl.load_workbook(filename)
    ws1 = source.active
    out = openpyxl.load_workbook(output)
    ws2 = out.active
    AnsDict = {'exp1_q5_pp':[[],[],[],[]],
               'exp1_pp1a':[[],[],[],[]],
               'Count_Target_In_Range_Order': [[],[],[],[]],
               'Total_Dict_Values_PP': [[],[],[],[]],
               'exp1_pp3':[[],[],[],[]]}
    AnsDict = collect_ans(ws1, AnsDict)
    questions = list(AnsDict.keys())
    if filt == []:
        data(filename, output, AnsDict, ws1, ws2, out, questions, dismode, distV)
    else:
        data(filename, output, AnsDict, ws1, ws2, out, filt, dismode, distV)
    df = pd.DataFrame
    
        


    
def main(source, output):
    cwd = os.getcwd()
    source_path = os.path.join(cwd, source)
    output_path = os.path.join(cwd, output)


if __name__ == "__main__":
    args = sys.argv
    print(args)
    if len(args) < 3:
        raise Exception("Please pass a source and target directory")
    source, output = args[1:]
    main(source, output)



    


        
    
    
