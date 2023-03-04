import openpyxl
import pandas
import os
import numpy as np
from itertools import chain
from openpyxl import Workbook

global curr_r

global distractor_mode


# check line 291915, 280421

def fill_in(ws1, ws2, start_r, end_r, ref_r, attempt):
    global curr_r
    #copy sid(2)>student_id(1), timestamp(3) > 2, div_id(6) > 8
    for r in range(start_r, end_r):
        sid = ws1.cell(row = ref_r, column = 2)
        ws2.cell(row = r, column = 1).value = sid.value
        time = ws1.cell(row = ref_r, column = 3)
        ws2.cell(row = r, column = 2).value = time.value
        probname = ws1.cell(row = ref_r, column = 6)
        ws2.cell(row = r, column = 7).value = probname.value
        ws2.cell(row = r, column = 3).value ="submit solution"
        ws2.cell(row = r, column = 4).value ="parsons"
        ws2.cell(row = r, column = 5).value = attempt
        # is last attempt
        ws2.cell(row = r, column = 6).value = 0
        if(ws1.cell(row = ref_r+1, column = 6).value != ws1.cell(row = ref_r, column = 6).value):
            ws2.cell(row = r, column = 6).value = 1
    
    curr_r = end_r


def collect_ans(ws1,AnsDict,mr):

    #answer
    (AnsDict['exp1_q5_pp'])[0] = ['0_0','1_1','2_1','3_4_2','7_2','9_2','10_1']
    (AnsDict['exp1_pp1a'])[0] = ['0_0','1_1','3_1','4_2','6_3','8_2','9_1']
    (AnsDict['Count_Target_In_Range_Order'])[0] = ['0_0','1_1','3_1','5_2','7_2','9_3','11_1']
    (AnsDict['Total_Dict_Values_PP'])[0] = ['0_0','2_1','3_1','5_2','7_1']
    (AnsDict['exp1_pp3'])[0] = ['0_0','2_1','3_1','4_1']
    
    # distractors
    (AnsDict['exp1_q5_pp'])[1] = ['5','6','8']
    (AnsDict['exp1_pp1a'])[1] = ['2','5','7']
    (AnsDict['Count_Target_In_Range_Order'])[1] = ['0','1','2', '4', '6', '8','10']
    (AnsDict['Total_Dict_Values_PP'])[1] = ['1','4', '6']
    (AnsDict['exp1_pp3'])[1] = ['1','6']

    #KC (if no corresponding kc, use '' as fill in
    #Attribute: 
    (AnsDict['exp1_q5_pp'])[2] = ['FunctionDef,args','Assign','For','Assign,get() function,Attribute,args','Assign,get() function,Attribute,args','Assign,BinOp,Add','append() function,Attribute,args','Return']
    (AnsDict['exp1_pp1a'])[2] = ['FunctionDef, args','Assign','While, Comparison Operator < Less than, len() function, args','If Statement, BoolOp, And, Comparison Operator == Equal, Subscript','Return','Assignment Operator +=','Return']
    (AnsDict['Count_Target_In_Range_Order'])[2] = ['FunctionDef,args','Assign','For, range() function, args,BinOp, Add','Assign, Subscript','If Statement,Comparison Operator == Equal','Assign,BinOp,Add','Return']
    (AnsDict['Total_Dict_Values_PP'])[2] = ['FunctionDef,args','Assign','For','Assignment Operator +=, Subscript','Return']
    (AnsDict['exp1_pp3'])[2] = ['FunctionDef,args','Assign, max() function, args','Assign, max() function, args','Return,BinOp, Sub']
    
    return AnsDict

    


def find_changes(last, curr, probname, line):
##    line = last + 1
##    while (str(ws1.cell(row = i, column = 6).value) == probname) and (str(ws1.cell(row = i, column = 4).value) == 'parsonsMove') and line < curr:
##        content = ws1.cell(row = i, column = 5).value
##        typ,le,ri,c = content.split('|')
##        if typ == 'move':
    changes = []
    deleted = []
    res = []
    if curr[0] == '':
        return []
    if last[0] == '':
        return curr
    curr_tmp = []
    curr_blocks = []
    # 1. find new added elements
    for item in curr:
        if item not in last:
            changes.append(item)
        curr_tmp.append(item)
        curr_blocks.append((item.split('_'))[0])
            


    # fill in places for removed blocks so to compare change in order for existing blocks
    while len(curr_tmp) < len(last):
        for i in range(0, len(last)):
            if ((last[i]).split('_'))[0] not in curr_blocks:
                curr_tmp.insert(i, 'block removed')
                
    #  find elements that changed order
    for block in curr_tmp:
        if block != 'block removed' and block in last and (curr_tmp.index(block) != last.index(block)):
            changes.append(block)
    # 5. iterate through curr, if the item is in the change list, put them in return change in order
    for item in curr:
        if item in changes:
            res.append(item)
    return res
    

def get_attempt(c,line,ws1):
    #find attempt num
    # check 289126
    res = c.split('-')
    apt = ''
    attempt = 0
    for letter in res:
        if letter[0] == 'c':
            apt = letter[1:len(letter)]
    attempt = int(apt)
    if 's' not in res:
        return attempt
    apt = ''
    attempt = 0
    # if the users reset the attempts, find the actual attempt # from last parsonsmove line
    if 's' in res:
        l = line-1
        while str(ws1.cell(row = l, column = 4).value) != 'parsonsMove' and str(ws1.cell(row = l, column = 4).value) != 'parsons':
            l = l-1
        if line == 254503:
            print(l)
        content_prev = ws1.cell(row = l, column = 5).value
        typ2,le2,ri2,c2 = content_prev.split('|')
        res2 = c2.split('-')
        for letters in res2:
            if letters[0] == 'c':
                apt = letters[1:len(letters)]
        attempt = int(apt) + 1
    return attempt


def add_distractor(ans, curr, distractor):
    #check line 280421
    # 1. find the block (no indentation) which is in distractor
    res = []
    for item in ans:
        res.append(item)
    for block in curr:
        if (block.split('_'))[0] in distractor:
            # insert at that index of curr_ans_tmp
            if distractor_mode == True:
                res.insert(curr.index(block), "distractor")
            else:
                res.append('extra')
            
    # return curr_ans_tmp
    return res



def add_KC(initial, ans_tmp):
    res = []
    for item in initial:
        res.append(item)
    for string in ans_tmp:
        if string == 'distractor' and distractor_mode == True:
            res.insert(ans_tmp.index(string), '')
        if string == 'extra' and distractor_mode == False:
            res.append('')
    return res
            


    
def data(filename, output, dismode=True):
    source = openpyxl.load_workbook(filename)
    ws1 = source.active
    out = openpyxl.load_workbook(output)
    ws2 = out.active
    wb = Workbook()
    tmp = wb.active
    global distractor_mode
    distractor_mode = dismode
    

    #find someway to filter ws1
    filter_id= ['Total_Dict_Values_PP', 'exp1_q5_pp','exp1_pp1a','Count_Target_In_Range_Order']
    ws1.auto_filter.ref = "A0:J297132" 
    ws1.auto_filter.add_filter_column(6, filter_id);
    ws1.auto_filter.add_filter_column(4, ['parsons']);
    #ws1 = ws1[((ws1['div_id'] == 'exp1_q5_pp') | (ws1['div_id'] == 'exp1_pp1a') | (ws1['div_id'] == 'Count_Target_In_Range_Order')|(ws1['div_id'] == 'Total_Dict_Values_PP'))&(ws1['event'] == 'parsons')]



    AnsDict = {'exp1_q5_pp':[[],[],[]],
               'exp1_pp1a':[[],[],[]],
               'Count_Target_In_Range_Order': [[],[],[]],
               'Total_Dict_Values_PP': [[],[],[]],
               'exp1_pp3':[[],[],[]]}

               

    mr = ws1.max_row
    mc = ws1.max_column
    ws2["A1"] = "student_Id" 
    ws2["B1"] = "timestamp"
    ws2["C1"] = "student response type" 
    ws2["D1"] = "problem type" 
    ws2["E1"] = "attempt at problem type"
    ws2["F1"] = "is last attempt"   
    ws2["G1"] = "problem name"
    ws2["H1"] = "outcome"  #correct / incorrect
    ws2["I1"] = "input"  #?the block dragged in right
    ws2["J1"] = "KC" #make a list to correspond for each step?
    ws2["K1"] = "reference line"
    ws2["L1"] = "should be"

    AnsDict = collect_ans(ws1, AnsDict, mr)
                
    r = 2;
    last_attempt = []
    correct = False;


    questions = list(AnsDict.keys())

    
    for i in range(2, mr+1):
        if (str(ws1.cell(row = i, column = 4).value) == 'parsons') and ((str(ws1.cell(row = i, column = 6).value) == 'Total_Dict_Values_PP')
                        or (str(ws1.cell(row = i, column = 6).value) == 'exp1_q5_pp')
                        or (str(ws1.cell(row = i, column = 6).value) == 'exp1_pp1a')
                        or (str(ws1.cell(row = i, column = 6).value) == 'Count_Target_In_Range_Order')
                        or (str(ws1.cell(row = i, column = 6).value) == 'exp1_pp3')):
            probname = ws1.cell(row = i, column = 6).value
            content = ws1.cell(row = i, column = 5).value
            typ,le,ri,c = content.split('|')

            #get attempt number
            attempt = 0;
            attempt = get_attempt(c,i,ws1)
            ws2.cell(row = r, column = 5).value = attempt
            
            #check correctness
            if typ == 'incorrect':
                ws2.cell(row = r, column = 8).value = 'INCORRECT'
                ws2.cell(row = r, column = 11).value = i
                correct = False
            
            if typ == 'correct':
                ws2.cell(row = r, column = 8).value = 'CORRECT'
                correct = True
                ws2.cell(row = r, column = 11).value = i


            #look into the input result
            #ws2.cell(row = r, column = 11).value = ri
            inp = ri.split('-')
            correct_ans = AnsDict[probname][0]
            correct_ans_tmp = []
            distractors = AnsDict[probname][1]
            KC = AnsDict[probname][2]
            KC_tmp = []


            if (len(inp) > len(correct_ans)):
                correct_ans_tmp = add_distractor(correct_ans, inp, (AnsDict[probname])[1])
                KC_tmp = add_KC(KC,correct_ans_tmp)
            else:
                for item in correct_ans:
                    correct_ans_tmp.append(item)
                for item in KC:
                    KC_tmp.append(item)

                        

            count_change = 0            
            #when it's first submission
            if(last_attempt == []):
                for x in range(0, len(inp)):
                    if(correct):#if correct, make len(inp) rows and fill 8th col in with correct
                        ws2.cell(row = r + x, column = 8).value = 'CORRECT'
                        ws2.cell(row = r + x, column = 9).value = inp[x]
                        ws2.cell(row = r + x, column = 10).value = KC[x]
                    else:
                        #check all code blocks and find those incorrect 
                        if (inp[x] != correct_ans_tmp[x]) and ((correct_ans_tmp[x] == 'distractor') or (correct_ans_tmp[x] == 'extra')):
                            ws2.cell(row = r + x, column = 8).value = 'INCORRECT'
                            ws2.cell(row = r + x, column = 9).value = inp[x]
                            ws2.cell(row = r + x, column = 12).value = correct_ans_tmp[x]
                        if (inp[x] != correct_ans_tmp[x]) and (correct_ans_tmp[x] != 'distractor') and (correct_ans_tmp[x] != 'extra'):
                            # if the distractor is replace of a correct block (absolute
                            if (inp[x] in correct_ans) and (correct_ans.index(inp[x]) == x):
                                ws2.cell(row = r + x, column = 8).value = 'CORRECT'
                                ws2.cell(row = r + x, column = 9).value = inp[x]
                                ws2.cell(row = r + x, column = 10).value = KC[x]
                            else:
                                ws2.cell(row = r + x, column = 8).value = 'INCORRECT'
                                ws2.cell(row = r + x, column = 9).value = inp[x]
                                ws2.cell(row = r + x, column = 12).value = correct_ans_tmp[x]
                                ws2.cell(row = r + x, column = 10).value = KC_tmp[x]
        
                        #if the distractor is an added extra block to the answer(relative order)
                        if (inp[x] == correct_ans_tmp[x]):
                            ws2.cell(row = r + x, column = 8).value = 'CORRECT'
                            ws2.cell(row = r + x, column = 9).value = inp[x]
                            ws2.cell(row = r + x, column = 10).value = KC_tmp[x]
                    count_change = len(inp)
            
            #if not first attempt, compare with last attempt to find the different ones, check their correctness
            if last_attempt != []:
                # find out moves between line of last attempt and current i th line, put in a list to iterate through to check correctness
                changed = find_changes(last_attempt, inp, probname, i)
                count_change = len(changed)
                if changed == []:
                    count_change = 1

                
                for y in range(0, len(changed)):
                    idx = inp.index(changed[y])
                    if (inp[idx] != correct_ans_tmp[idx]) and ((correct_ans_tmp[idx] == 'distractor')or (correct_ans_tmp[idx] == 'extra')):
                        ws2.cell(row = r + y, column = 8).value = 'INCORRECT'
                        ws2.cell(row = r + y, column = 9).value = inp[idx]
                        ws2.cell(row = r + y, column = 12).value = correct_ans_tmp[idx]
                    if (inp[idx] != correct_ans_tmp[idx]) and (correct_ans_tmp[idx] != 'distractor') and (correct_ans_tmp[idx] != 'extra'):
                        if (inp[idx] in correct_ans) and (correct_ans.index(inp[idx]) == idx):
                            ws2.cell(row = r + y, column = 8).value = 'CORRECT'
                            ws2.cell(row = r + y, column = 9).value = inp[idx]
                            ws2.cell(row = r + y, column = 10).value = KC[idx]
                        else:
                            ws2.cell(row = r + y, column = 8).value = 'INCORRECT'
                            ws2.cell(row = r + y, column = 9).value = inp[idx]
                            ws2.cell(row = r + y, column = 12).value = correct_ans_tmp[idx]
                            ws2.cell(row = r + y, column = 10).value = KC_tmp[idx]
                    if (inp[idx] == correct_ans_tmp[idx]):
                        ws2.cell(row = r + y, column = 8).value = 'CORRECT'
                        ws2.cell(row = r + y, column = 9).value = inp[idx]
                        ws2.cell(row = r + y, column = 10).value = KC_tmp[idx]

                    
            last_attempt = inp;
            end_r = r + count_change
            fill_in(ws1, ws2, r, end_r, i, attempt)
            r = end_r
            if(ws1.cell(row = i+1, column = 6).value != ws1.cell(row = i, column = 6).value):
                last_attempt = []


    
    out.save(str(output))



    
def main(source, output):
    cwd = os.getcwd()
    source_path = os.path.join(cwd, source)
    output_path = os.path.join(cwd, output)


if __name__ == "__main__":
    args = sys.argv
    print(args)
    if len(args) < 3:
        raise Exception("Please pass a source and target directory")
    source, output = args[1:]
    main(source, output)



    


        
    
    
