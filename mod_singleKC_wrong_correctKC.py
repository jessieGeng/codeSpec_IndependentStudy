import openpyxl
# import pandas
import os
import numpy as np
from itertools import chain
from openpyxl import Workbook
import pandas as pd

# global curr_r

global distractor_mode

global questions;

global empty;


# check line 291915, 280421

def fill_in(ws1, ws2, start_r, end_r, ref_r, attempt):
    #copy sid(2)>student_id(1), timestamp(3) > 2, div_id(6) > 8
    for r in range(start_r, end_r):
        sid = ws1.cell(row = ref_r, column = 2)
        ws2.cell(row = r, column = 1).value = sid.value
        time = ws1.cell(row = ref_r, column = 3)
        ws2.cell(row = r, column = 2).value = time.value
        probname = ws1.cell(row = ref_r, column = 6)
        ws2.cell(row = r, column = 9).value = probname.value
        ws2.cell(row = r, column = 3).value ="ATTEMPT"
        ws2.cell(row = r, column = 4).value ="Parsons problems"
        ws2.cell(row = r, column = 5).value = "programming"
        ws2.cell(row = r, column = 7).value = attempt
        ws2.cell(row = r, column = 13).value = 1
        # is last attempt
        ws2.cell(row = r, column = 8).value = 0
        if(ws1.cell(row = ref_r+1, column = 6).value != ws1.cell(row = ref_r, column = 6).value):
            ws2.cell(row = r, column = 8).value = 1

    

def collect_ans(ws1,AnsDict):

    #answer
    (AnsDict['exp1_q5_pp'])[0] = ['0_0','1_1','2_1','3_4_2','7_2','9_2','10_1']
    (AnsDict['exp1_pp1a'])[0] = ['0_0','1_1','3_1','4_2','6_3','8_2','9_1']
    (AnsDict['Count_Target_In_Range_Order'])[0] = ['0_0','1_1','3_1','5_2','7_2','9_3','11_1']
    (AnsDict['Total_Dict_Values_PP'])[0] = ['0_0','2_1','3_1','5_2','7_1']
    (AnsDict['exp1_pp3'])[0] = ['0_0','2_1','3_1','4_1']
    
    # distractors
    (AnsDict['exp1_q5_pp'])[1] = ['5','6','8']
    (AnsDict['exp1_pp1a'])[1] = ['2','5','7']
    (AnsDict['Count_Target_In_Range_Order'])[1] = ['2', '4', '6', '8','10']
    (AnsDict['Total_Dict_Values_PP'])[1] = ['1','4', '6']
    (AnsDict['exp1_pp3'])[1] = ['1','6']

    #KC (if no corresponding kc, use '' as fill in
    #Attribute: 
    (AnsDict['exp1_q5_pp'])[2] = ['FunctionDef,args','Assign','For','Assign,get() function,Attribute,args','Assign,get() function,Attribute,args','Assign,BinOp,Add','append() function,Attribute,args','Return']
    (AnsDict['exp1_pp1a'])[2] = ['FunctionDef,args','Assign','While,Comparison Operator < Less than,len() function,args','If Statement,BoolOp,And,Comparison Operator == Equal,Subscript','Return','Assignment Operator +=','Return']
    (AnsDict['Count_Target_In_Range_Order'])[2] = ['FunctionDef,args','Assign','For,range() function,args,BinOp,Add','Assign,Subscript','If Statement,Comparison Operator == Equal','Assign,BinOp,Add','Return']
    (AnsDict['Total_Dict_Values_PP'])[2] = ['FunctionDef,args','Assign','For','Assignment Operator +=,Subscript','Return']
    (AnsDict['exp1_pp3'])[2] = ['FunctionDef,args','Assign,max() function,args','Assign,max() function,args','Return,BinOp,Sub']

    # distractor KC
    (AnsDict['exp1_q5_pp'])[3] = ['Assign,get() function,Attribute,args','Assign,get() function,Attribute,args','Assign,BinOp,Add']
    (AnsDict['exp1_pp1a'])[3] = ['Assign','If Statement,BoolOp,And,Comparison Operator == Equal,Subscript','Return']
    (AnsDict['Count_Target_In_Range_Order'])[3] = ['Assign', 'For,range() function,args,BinOp,Add', 'Assign,Subscript', 'If Statement,Comparison Operator == Equal','Assign,BinOp,Add']
    (AnsDict['Total_Dict_Values_PP'])[3] = ['FunctionDef,args','For', 'Assignment Operator +=, Subscript']
    (AnsDict['exp1_pp3'])[3] = ['FunctionDef,args','Return,BinOp, Sub']
    
    return AnsDict

    
    

def find_changes(last, curr, probname, line):
    changes = []
    deleted = []
    res = []
    if curr[0] == '':
        return []
    if last[0] == '':
        return curr
    curr_tmp = []
    curr_blocks = []
    # 1. find new added elements
    for item in curr:
        if item not in last:
            changes.append(item)
        curr_tmp.append(item)
        curr_blocks.append((item.split('_'))[0])
            


    # fill in places for removed blocks so to compare change in order for existing blocks
    while len(curr_tmp) < len(last):
        for i in range(0, len(last)):
            if ((last[i]).split('_'))[0] not in curr_blocks:
                curr_tmp.insert(i, 'block removed')
                
    #  find elements that changed order
    for block in curr_tmp:
        if block != 'block removed' and block in last and (curr_tmp.index(block) != last.index(block)):
            changes.append(block)
    # 5. iterate through curr, if the item is in the change list, put them in return change in order
    for item in curr:
        if item in changes:
            res.append(item)
    return res
    

def get_attempt(c,line,ws1):
    #find attempt num
    # check 289126
    res = c.split('-')
    apt = ''
    attempt = 0
    for letter in res:
        if letter[0] == 'c':
            apt = letter[1:len(letter)]
    attempt = int(apt)
    if 's' not in res:
        return attempt
    apt = ''
    attempt = 0
    # if the users reset the attempts, find the actual attempt # from last parsonsmove line
    if 's' in res:
        l = line-1
        while str(ws1.cell(row = l, column = 4).value) != 'parsonsMove' and str(ws1.cell(row = l, column = 4).value) != 'parsons':
            l = l-1
        content_prev = ws1.cell(row = l, column = 5).value
        typ2,le2,ri2,c2 = content_prev.split('|')
        res2 = c2.split('-')
        for letters in res2:
            if letters[0] == 'c':
                apt = letters[1:len(letters)]
        attempt = int(apt) + 1
    return attempt


def add_distractor(ans, curr, distractor):
    # 1. find the block (no indentation) which is in distractor
    res = []
    for item in ans:
        res.append(item)
    for block in curr:
        bkstart = (block.split('_'))[0]
        if bkstart in distractor:
            # insert at that index of curr_ans_tmp
            if distractor_mode == True:
                addinto = "distractor " + str(bkstart)
                res.insert(curr.index(block), addinto)
            else:
                addinto = "extra " + str(bkstart)
                res.append(addinto)
            
    # return curr_ans_tmp
    return res



def add_KC(initial, ans_tmp, KC_distractor, distractorlist):
    res = []
    for item in initial:
        res.append(item)
    for string in ans_tmp:
        if string[0] == 'd' and distractor_mode == True:
            distractor = (string.split())[1]
            inx = distractorlist.index(distractor)
            res.insert(ans_tmp.index(string), KC_distractor[inx])
        if string[0] == 'e' and distractor_mode == False:
            distractor = (string.split())[1]
            inx = distractorlist.index(distractor)
            res.append(KC_distractor[inx])
    return res
            

def split_KCs(start_r, end_r, ws2, KClist, dist):
    for i in range(start_r, end_r):
        ws2.cell(row = i, column = 10).value = ws2.cell(row = start_r, column = 10).value
        ws2.cell(row = i, column = 11).value = ws2.cell(row = start_r, column = 11).value
        if dist:
            ws2.cell(row = i, column = 6).value = "distractor " + str(ws2.cell(row = start_r, column = 11).value)
        else:
            ws2.cell(row = i, column = 6).value = KClist[i-start_r]
        ws2.cell(row = i, column = 12).value = KClist[i-start_r]
    
    
    
def data(filename, output, AnsDict, ws1, ws2, out, filt, dismode=True):
    global distractor_mode
    distractor_mode = dismode
               

    mr = ws1.max_row
    mc = ws1.max_column
    ws2["A1"] = "Anon Student Id" 
    ws2["B1"] = "Time"
    ws2["C1"] = "Student Response Type" 
    ws2["D1"] = "Student Response Subtype"
    ws2["E1"] = "Level(default)"
    ws2["F1"] = "Step Name"
    ws2["G1"] = "Attempt At Step"
    ws2["H1"] = "Is Last Attempt"   
    ws2["I1"] = "Problem Name"
    ws2["J1"] = "Outcome"  #correct / incorrect
    ws2["K1"] = "Input"  #?the block dragged in right
    ws2["L1"] = "KC(Tokens)" #make a list to correspond for each step?
    ws2["M1"] = "Session Id"
    ws2["N1"] = "reference line"
    ws2["O1"] = "should be"

                
    r = 2;
    last_attempt = []
    correct = False;

        

    
    for i in range(2, mr+1):
         if(str(ws1.cell(row = i, column = 4).value) == 'parsons') and ((str(ws1.cell(row = i, column = 6).value) in filt)):
            probname = ws1.cell(row = i, column = 6).value
            content = ws1.cell(row = i, column = 5).value
            typ,le,ri,c = content.split('|')

            #get attempt number
            attempt = 0;
            attempt = get_attempt(c,i,ws1)
            ws2.cell(row = r, column = 7).value = attempt
            
            #check correctness
            if typ == 'incorrect':
                ws2.cell(row = r, column = 10).value = 'INCORRECT'
                ws2.cell(row = r, column = 14).value = i
                correct = False
            
            if typ == 'correct':
                ws2.cell(row = r, column = 10).value = 'CORRECT'
                correct = True
                ws2.cell(row = r, column = 14).value = i


            #look into the input result
            inp = ri.split('-')
            correct_ans = AnsDict[probname][0]
            correct_ans_tmp = []
            distractors = AnsDict[probname][1]
            KC = AnsDict[probname][2]
            KC_distractor = AnsDict[probname][3]
            KC_tmp = []


            if (len(inp) > len(correct_ans)):
                correct_ans_tmp = add_distractor(correct_ans, inp, (AnsDict[probname])[1])
                KC_tmp = add_KC(KC,correct_ans_tmp,KC_distractor, distractors)
            else:
                for item in correct_ans:
                    correct_ans_tmp.append(item)
                for item in KC:
                    KC_tmp.append(item)

                        

            count_change = 0            
            #when it's first submission
            if(last_attempt == []):
                for x in range(0, len(inp)):
                    if(correct):#if correct, make len(inp) rows and fill 8th col in with correct
                        ws2.cell(row = r, column = 10).value = 'CORRECT'
                        ws2.cell(row = r, column = 11).value = inp[x]
                        KClist = (KC[x]).split(",")
                        count_change = count_change + len(KClist)
                        split_KCs(r, r + len(KClist), ws2, KClist, False)
                        r = r + len(KClist)
                    else:
                        #check all code blocks and find those incorrect 
                        if (inp[x] != correct_ans_tmp[x]) and (((correct_ans_tmp[x])[0] == 'd') or ((correct_ans_tmp[x])[0] == 'e')):
                            ws2.cell(row = r, column = 10).value = 'INCORRECT'
                            ws2.cell(row = r, column = 11).value = inp[x]
                            ws2.cell(row = r, column = 15).value = correct_ans_tmp[x]
                            KClist = (KC_tmp[x]).split(",")
                            count_change = count_change + len(KClist)
                            split_KCs(r, r + len(KClist), ws2, KClist, True)
                            r = r + len(KClist)
                        if (inp[x] != correct_ans_tmp[x]) and ((correct_ans_tmp[x])[0] != 'd') and ((correct_ans_tmp[x])[0] != 'e'):
                            # if the distractor is replace of a correct block (absolute
                            if (inp[x] in correct_ans) and (correct_ans.index(inp[x]) == x):
                                ws2.cell(row = r, column = 10).value = 'CORRECT'
                                ws2.cell(row = r, column = 11).value = inp[x]
                                KClist = (KC[x]).split(",")
                                count_change = count_change + len(KClist)
                                split_KCs(r, r + len(KClist), ws2, KClist, False)
                                r = r + len(KClist)
                            else:
                                ws2.cell(row = r, column = 10).value = 'INCORRECT'
                                ws2.cell(row = r, column = 11).value = inp[x]
                                ws2.cell(row = r, column = 15).value = correct_ans_tmp[x]
                                KClist = (KC_tmp[x]).split(",")
                                count_change = count_change + len(KClist)
                                split_KCs(r, r + len(KClist), ws2, KClist, False)
                                r = r + len(KClist)
        
                        #if the distractor is an added extra block to the answer(relative order)
                        if (inp[x] == correct_ans_tmp[x]):
                            ws2.cell(row = r, column = 10).value = 'CORRECT'
                            ws2.cell(row = r, column = 11).value = inp[x]
                            ws2.cell(row = r, column = 6).value = KC_tmp[x]
                            ws2.cell(row = r, column = 12).value = KC_tmp[x]
                            KClist = (KC_tmp[x]).split(",")
                            count_change = count_change + len(KClist)
                            split_KCs(r, r + len(KClist), ws2, KClist, False)
                            r = r + len(KClist)
                    
            
            #if not first attempt, compare with last attempt to find the different ones, check their correctness
            if last_attempt != []:
                # find out moves between line of last attempt and current i th line, put in a list to iterate through to check correctness
                changed = find_changes(last_attempt, inp, probname, i)
                count_change = len(changed)
                if changed == []:
                    count_change = 1
                    ws2.cell(row = r, column = 6).value = probname + " empty/unchanged in attempt " + str(attempt)
                    r = r + 1

                
                for y in range(0, len(changed)):
                    idx = inp.index(changed[y])
                    if (inp[idx] != correct_ans_tmp[idx]) and (((correct_ans_tmp[idx])[0] == 'd')or ((correct_ans_tmp[idx])[0] == 'e')):
                        ws2.cell(row = r, column = 10).value = 'INCORRECT'
                        ws2.cell(row = r, column = 11).value = inp[idx]
                        ws2.cell(row = r, column = 15).value = correct_ans_tmp[idx]
                        KClist = (KC_tmp[idx]).split(",")
                        count_change = count_change + len(KClist)
                        split_KCs(r, r + len(KClist), ws2, KClist, True)
                        r = r + len(KClist)
                        
                    if (inp[idx] != correct_ans_tmp[idx]) and ((correct_ans_tmp[idx])[0] != 'd') and ((correct_ans_tmp[idx])[0] != 'e'):
                        if (inp[idx] in correct_ans) and (correct_ans.index(inp[idx]) == idx):
                            ws2.cell(row = r, column = 10).value = 'CORRECT'
                            ws2.cell(row = r, column = 11).value = inp[idx]
                            KClist = (KC[idx]).split(",")
                            count_change = count_change + len(KClist)
                            split_KCs(r, r + len(KClist), ws2, KClist, False)
                            r = r + len(KClist)
                        else:
                            ws2.cell(row = r, column = 10).value = 'INCORRECT'
                            ws2.cell(row = r, column = 11).value = inp[idx]
                            ws2.cell(row = r, column = 15).value = correct_ans_tmp[idx]
                            KClist = (KC_tmp[idx]).split(",")
                            count_change = count_change + len(KClist)
                            split_KCs(r, r + len(KClist), ws2, KClist, False)
                            r = r + len(KClist)
                    if (inp[idx] == correct_ans_tmp[idx]):
                        ws2.cell(row = r, column = 10).value = 'CORRECT'
                        ws2.cell(row = r, column = 11).value = inp[idx]
                        KClist = (KC_tmp[idx]).split(",")
                        count_change = count_change + len(KClist)
                        split_KCs(r, r + len(KClist), ws2, KClist, False)
                        r = r + len(KClist)

                    
            last_attempt = inp;
            start_r = r - count_change
            fill_in(ws1, ws2, start_r, r, i, attempt)
            if(ws1.cell(row = i+1, column = 6).value != ws1.cell(row = i, column = 6).value):
                last_attempt = []

    
    out.save(str(output))




def init(filename, output, filt=[], dismode=True):
    source = openpyxl.load_workbook(filename)
    ws1 = source.active
    out = openpyxl.load_workbook(output)
    ws2 = out.active
    AnsDict = {'exp1_q5_pp':[[],[],[],[]],
               'exp1_pp1a':[[],[],[],[]],
               'Count_Target_In_Range_Order': [[],[],[],[]],
               'Total_Dict_Values_PP': [[],[],[],[]],
               'exp1_pp3':[[],[],[],[]]}
    AnsDict = collect_ans(ws1, AnsDict)
    questions = list(AnsDict.keys())
    if filt == []:
        data(filename, output, AnsDict, ws1, ws2, out, questions, dismode)
    else:
        data(filename, output, AnsDict, ws1, ws2, out, filt, dismode)
    df = pd.DataFrame
    
        


    
def main(source, output):
    cwd = os.getcwd()
    source_path = os.path.join(cwd, source)
    output_path = os.path.join(cwd, output)


if __name__ == "__main__":
    args = sys.argv
    print(args)
    if len(args) < 3:
        raise Exception("Please pass a source and target directory")
    source, output = args[1:]
    main(source, output)



    


        
    
    
